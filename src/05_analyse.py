"""Stage 05 — Analyse intervention results.

Loads the unified intervention results from Stage 04 and produces:

  - Per-idiom output-space intervention slopes
  - Layer sensitivity matrix from activation-space interventions
  - Output intervention plot (base-normalised) with:
      · robustness check event markers (vertical lines, colour-coded by severity)
      · pre-active-era shading (grey) for each idiom's currency introduction
      · first-attestation tick marks per idiom
  - Activation intervention plot (base-normalised, best layer)
  - Layer sensitivity heatmap

Outputs
-------
data/results/intervention_slopes.csv
data/results/layer_sensitivity.npy
data/results/layer_sensitivity_index.csv
data/plots/output_intervention.png
data/plots/activation_intervention.png
data/plots/layer_heatmap.png

Usage
-----
    python src/05_analyse.py
    python src/05_analyse.py --dry-run
"""

from __future__ import annotations

import argparse
import logging
import sys
import warnings
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
import seaborn as sns
import yaml
from sklearn.linear_model import LinearRegression

try:
    import openpyxl  # noqa: F401 — only needed for real-value plot
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from config_utils import get_model_cfg, resolve_paths

CONFIG_PATH = PROJECT_ROOT / "config" / "config.yaml"

logger = logging.getLogger(__name__)

# Severity → line style for robustness check markers
_SEVERITY_STYLE = {
    "high":   {"color": "#d62728", "linestyle": "--",  "alpha": 0.55, "linewidth": 1.2},
    "medium": {"color": "#ff7f0e", "linestyle": "-.",  "alpha": 0.45, "linewidth": 0.9},
    "low":    {"color": "#aec7e8", "linestyle": ":",   "alpha": 0.40, "linewidth": 0.8},
}


def _normalise_to_base(values: np.ndarray) -> np.ndarray:
    """Divide a score series by its first element so it starts at 1.0.

    If the first value is near zero the series is returned unchanged to avoid
    division explosions.
    """
    first = values[0]
    if abs(first) < 1e-9:
        return values
    return values / first


def _ols_slope(x: np.ndarray, y: np.ndarray) -> tuple[float, float]:
    """Return (slope, slope_se) from simple OLS of y ~ x."""
    n = len(x)
    if n < 3:
        return float("nan"), float("nan")
    reg = LinearRegression(fit_intercept=True).fit(x.reshape(-1, 1), y)
    slope = float(reg.coef_[0])
    residuals = y - reg.predict(x.reshape(-1, 1)).ravel()
    var_res = np.sum(residuals ** 2) / (n - 2)
    var_x   = np.sum((x - x.mean()) ** 2)
    se = float(np.sqrt(var_res / var_x)) if var_x > 0 else float("nan")
    return slope, se


def _plot_robustness_markers(
    ax: plt.Axes,
    checks: list[dict],
    x_min: float,
    x_max: float,
    y_top: float,
) -> None:
    """Draw vertical lines + rotated labels for robustness check events."""
    label_y  = y_top
    label_step = (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.04

    # Track label x-positions to stagger overlapping labels
    placed: list[float] = []

    for chk in checks:
        year = chk["year"]
        if not (x_min <= year <= x_max):
            continue
        style = _SEVERITY_STYLE.get(chk.get("severity", "low"), _SEVERITY_STYLE["low"])
        ax.axvline(year, **style)

        # Stagger label vertically if too close to previous
        lbl_y = y_top
        for px in placed:
            if abs(year - px) < (x_max - x_min) * 0.04:
                lbl_y -= label_step * 2
                break
        placed.append(year)

        ax.text(
            year + (x_max - x_min) * 0.004, lbl_y,
            chk["event"],
            rotation=90, va="top", ha="left",
            fontsize=5.5, color=style["color"], alpha=0.85,
        )


def _plot_real_value(
    xlsx_path: Path,
    cfg: dict,
    plots_dir: Path,
    x_min: float,
    x_max: float,
) -> None:
    """Load CPI and real-earnings series from the BoE millennium dataset and
    produce a two-panel time-series chart (``real_value.png``).

    Top panel   — Consumer Price Index (2015 = 100, log scale): shows the
                  general price level.  A rising line means inflation; a penny
                  buys proportionally less.

    Bottom panel — Real earnings index (1900 = 100, log scale): shows the
                   purchasing power of an average worker's wage.

    Both panels are annotated with denomination introduction / demonetisation
    windows (shaded bands) and robustness-check event lines.
    """
    if not _HAS_OPENPYXL:
        logger.warning("openpyxl not installed — skipping real_value.png")
        return
    if not xlsx_path.exists():
        logger.warning("Millennium data not found at %s — skipping real_value.png", xlsx_path)
        return

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # --- A47: CPI (col 3, 2015=100) ---
    ws47  = wb["A47. Wages and prices"]
    rows47 = list(ws47.iter_rows(values_only=True))
    cpi_years, cpi_vals = [], []
    for row in rows47[6:]:       # rows 0-5 are header/units
        yr, cpi = row[0], row[3]
        if isinstance(yr, int) and isinstance(cpi, float) and cpi > 0:
            if x_min <= yr <= x_max:
                cpi_years.append(yr)
                cpi_vals.append(cpi)

    # --- A48: Real earnings (col 1, 1900=100) ---
    ws48  = wb["A48. Real Earnings "]
    rows48 = list(ws48.iter_rows(values_only=True))
    re_years, re_vals = [], []
    for row in rows48[5:]:       # rows 0-4 are header/units
        yr, re = row[0], row[1]
        if isinstance(yr, int) and isinstance(re, float) and re > 0:
            if x_min <= yr <= x_max:
                re_years.append(yr)
                re_vals.append(re)

    wb.close()

    if not cpi_years:
        logger.warning("No CPI data in range [%g, %g] — skipping real_value.png", x_min, x_max)
        return

    cpi_arr = pd.Series(cpi_vals, index=cpi_years)
    re_arr  = pd.Series(re_vals,  index=re_years)

    checks = cfg.get("robustness_checks", [])
    denom_windows = cfg.get("denomination_windows", {})

    # Palette for denomination bands
    denom_colours = {
        "penny":    "#4878d0",
        "farthing": "#ee854a",
        "shilling": "#6acc65",
        "pound":    "#d65f5f",
    }

    def _draw_denom_bands(ax: plt.Axes) -> None:
        """Shade the active lifespan of each denomination."""
        for denom, info in denom_windows.items():
            intro = info.get("introduction_year", x_min)
            demo  = info.get("demonetisation_year", x_max)
            intro = max(intro, x_min)
            demo  = min(demo,  x_max)
            if intro >= demo:
                continue
            colour = denom_colours.get(denom, "grey")
            ax.axvspan(intro, demo, color=colour, alpha=0.07, zorder=0)

    def _draw_event_lines(ax: plt.Axes) -> None:
        for chk in checks:
            year = chk["year"]
            if not (x_min <= year <= x_max):
                continue
            style = _SEVERITY_STYLE.get(chk.get("severity", "low"), _SEVERITY_STYLE["low"])
            ax.axvline(year, zorder=2, **style)

    fig, (ax_top, ax_bot) = plt.subplots(
        2, 1, figsize=(14, 8), sharex=True,
        gridspec_kw={"hspace": 0.08},
    )

    # --- Top: CPI ---
    _draw_denom_bands(ax_top)
    ax_top.plot(cpi_arr.index, cpi_arr.values,
                color="#2c7bb6", linewidth=1.5, label="CPI (2015 = 100)")
    _draw_event_lines(ax_top)
    ax_top.set_yscale("log")
    ax_top.yaxis.set_major_formatter(mticker.ScalarFormatter())
    ax_top.set_ylabel("Price level (log, 2015 = 100)", fontsize=11)
    ax_top.set_title("UK price level and real earnings, 1200–1925\n"
                     "(shaded bands = denomination active era)", fontsize=12)
    ax_top.grid(axis="y", which="major", linestyle=":", alpha=0.4)

    # Denomination legend
    denom_patches = [
        mpatches.Patch(color=c, alpha=0.4, label=d.capitalize())
        for d, c in denom_colours.items()
        if d in denom_windows
    ]
    ax_top.legend(handles=denom_patches, loc="upper left",
                  fontsize=7, title="Denomination era", title_fontsize=7,
                  framealpha=0.6)

    # --- Bottom: Real earnings ---
    _draw_denom_bands(ax_bot)
    if re_arr.shape[0] > 0:
        ax_bot.plot(re_arr.index, re_arr.values,
                    color="#d7191c", linewidth=1.5, label="Real earnings (1900 = 100)")
    _draw_event_lines(ax_bot)
    ax_bot.set_yscale("log")
    ax_bot.yaxis.set_major_formatter(mticker.ScalarFormatter())
    ax_bot.set_ylabel("Real earnings (log, 1900 = 100)", fontsize=11)
    ax_bot.set_xlabel("Year", fontsize=11)
    ax_bot.grid(axis="y", which="major", linestyle=":", alpha=0.4)

    # Severity legend on bottom panel
    sev_patches = [
        mpatches.Patch(color=v["color"], alpha=0.7, label=f"{k.capitalize()} severity event")
        for k, v in _SEVERITY_STYLE.items()
    ]
    ax_bot.legend(handles=sev_patches, loc="upper left",
                  fontsize=7, title="Robustness checks", title_fontsize=7,
                  framealpha=0.6)

    ax_top.set_xlim(x_min, x_max)
    fig.tight_layout()
    out_path = plots_dir / "real_value.png"
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    logger.info("Saved → %s", out_path)


def run(cfg: dict, model_key: str, dry_run: bool) -> None:
    mcfg     = get_model_cfg(cfg, model_key)
    paths    = resolve_paths(cfg, model_key)
    n_layers = mcfg["n_layers"]
    checks   = cfg.get("robustness_checks", [])

    results_path = PROJECT_ROOT / paths["intervention_results"]
    slopes_path  = PROJECT_ROOT / paths["intervention_slopes"]
    mat_path     = PROJECT_ROOT / paths["layer_sensitivity_npy"]
    mat_idx_path = PROJECT_ROOT / paths["layer_sensitivity_idx"]
    plots_dir    = PROJECT_ROOT / paths["plots_dir"]
    plots_dir.mkdir(parents=True, exist_ok=True)

    # --- Load ---
    logger.info("Loading %s", results_path)
    df = pd.read_parquet(results_path)

    out_df = df[df["intervention_type"] == "output"].copy()
    act_df = df[df["intervention_type"] == "activation"].copy()

    if dry_run:
        idioms_sample = out_df["idiom"].unique()[:3]
        out_df = out_df[out_df["idiom"].isin(idioms_sample)]
        act_df = act_df[act_df["idiom"].isin(idioms_sample)]
        logger.info("[dry-run] Using idioms: %s", list(idioms_sample))

    idioms   = sorted(out_df["idiom"].unique())
    n_idioms = len(idioms)
    logger.info("Analysing %d idioms", n_idioms)

    # Per-idiom historical metadata (take first row per idiom)
    meta_cols = ["idiom", "denominations", "active_from", "first_attested"]
    available_meta = [c for c in meta_cols if c in out_df.columns]
    idiom_meta = (
        out_df[available_meta].drop_duplicates("idiom").set_index("idiom")
        if len(available_meta) > 1 else pd.DataFrame(index=idioms)
    )

    # --- Per-idiom output-space slopes ---
    slope_rows: list[dict] = []
    for idiom in idioms:
        sub = out_df[out_df["idiom"] == idiom]
        lam_arr   = sub["lambda_years"].values
        score_arr = sub["score_paraphrase"].values
        slope, se = _ols_slope(lam_arr, score_arr)
        baseline  = float(sub["score_baseline"].iloc[0])
        row: dict = {
            "idiom":                   idiom,
            "output_slope":            slope,
            "output_slope_se":         se,
            "baseline_score":          baseline,
            "output_slope_normalised": slope / abs(baseline) if abs(baseline) > 1e-10 else float("nan"),
        }
        # Attach metadata if available
        if idiom in idiom_meta.index:
            for col in ["denominations", "active_from", "first_attested"]:
                if col in idiom_meta.columns:
                    row[col] = idiom_meta.at[idiom, col]
        slope_rows.append(row)

    slopes_df = pd.DataFrame(slope_rows).sort_values("output_slope")
    slopes_df.to_csv(slopes_path, index=False)
    logger.info("Saved → %s", slopes_path)

    # --- Layer sensitivity matrix (activation-space) ---
    logger.info("Computing layer sensitivity matrix …")
    layer_mat = np.zeros((n_idioms, n_layers), dtype=np.float32)

    for i, idiom in enumerate(idioms):
        sub = act_df[act_df["idiom"] == idiom]
        for layer_idx in range(n_layers):
            layer_sub = sub[sub["intervention_layer"] == layer_idx + 1]
            if len(layer_sub) < 3:
                continue
            slope, _ = _ols_slope(
                layer_sub["lambda_years"].values,
                layer_sub["score_paraphrase"].values,
            )
            layer_mat[i, layer_idx] = slope

    np.save(mat_path, layer_mat)
    mat_idx_df = pd.DataFrame({"idiom": idioms})
    mat_idx_df.to_csv(mat_idx_path, index=False)
    logger.info("Saved layer sensitivity matrix → %s", mat_path)

    mean_abs_tmp = np.abs(layer_mat).mean(axis=0)
    # Exclude the final layer (index n_layers-1): injection at the last encoder
    # block's output passes through zero remaining transformer blocks, making it
    # mathematically identical to the output-space intervention.  Pick the best
    # layer among layers 1 … n_layers-1 (0-indexed: 0 … n_layers-2).
    best_layer   = int(np.argmax(mean_abs_tmp[:-1])) + 1   # 1-indexed, capped at n_layers-1

    # --- Print results ---
    print("\n=== Intervention Analysis Results ===\n")
    print(f"Layer with strongest temporal–semantic effect: Layer {best_layer}")
    print(f"  (mean |slope| across idioms = {mean_abs_tmp[best_layer-1]:.2e})\n")

    print("Per-idiom output slopes (Δtriviality per year of temporal shift):")
    display_cols = [c for c in [
        "idiom", "output_slope", "output_slope_se",
        "baseline_score", "output_slope_normalised",
        "active_from", "first_attested",
    ] if c in slopes_df.columns]
    print(slopes_df[display_cols].to_string(index=False))

    n_neg = (slopes_df["output_slope"] < 0).sum()
    n_pos = (slopes_df["output_slope"] > 0).sum()
    print(f"\n  Negative slopes (more modern → more significant): {n_neg}/{n_idioms}")
    print(f"  Positive slopes (more modern → more trivial):     {n_pos}/{n_idioms}")

    # Shared plot helpers
    x_min   = float(out_df["lambda_years"].min())
    x_max   = float(out_df["lambda_years"].max())
    palette = sns.color_palette("tab20", n_colors=n_idioms)

    global_active_from: int | None = None
    if "active_from" in idiom_meta.columns:
        global_active_from = int(idiom_meta["active_from"].min())

    def _draw_shared_annotations(ax: plt.Axes, x_min: float, x_max: float) -> None:
        """Shade pre-coin era, draw first-attestation ticks, horizontal zero line."""
        if global_active_from is not None and global_active_from > x_min:
            ax.axvspan(x_min, global_active_from, color="lightgrey", alpha=0.35)
        if "first_attested" in idiom_meta.columns:
            for i, idiom in enumerate(idioms):
                fa = idiom_meta.at[idiom, "first_attested"] if idiom in idiom_meta.index else None
                if fa and not pd.isna(fa) and x_min <= fa <= x_max:
                    ax.axvline(fa, color=palette[i], linewidth=0.6,
                               linestyle="-", alpha=0.3)
        ax.axhline(1.0, color="black", linewidth=0.8, linestyle="--", alpha=0.4)

    def _add_legends(ax: plt.Axes, checks: list[dict],
                     x_min: float, x_max: float) -> None:
        """Add idiom legend (left) and event legend (right)."""
        idiom_handles = [
            plt.Line2D([0], [0], color=palette[i], linewidth=1.5, label=idiom)
            for i, idiom in enumerate(idioms)
        ]
        legend_left = ax.legend(
            handles=idiom_handles,
            loc="upper left", fontsize=6.5, ncol=1,
            framealpha=0.55, title="Idioms", title_fontsize=7,
        )
        ax.add_artist(legend_left)

        event_patches = [
            mpatches.Patch(color=v["color"], alpha=0.7, label=f"Robustness: {k} severity")
            for k, v in _SEVERITY_STYLE.items()
        ]
        if global_active_from is not None and global_active_from > x_min:
            event_patches.append(
                mpatches.Patch(color="lightgrey", alpha=0.5, label="Pre-coin era")
            )
        ax.legend(
            handles=event_patches,
            loc="upper right", fontsize=7, framealpha=0.6,
            title="Events", title_fontsize=7,
        )

    # --- Plot 1: Output intervention curves (base-normalised) ---
    fig, ax = plt.subplots(figsize=(14, 7))
    _draw_shared_annotations(ax, x_min, x_max)

    for i, idiom in enumerate(idioms):
        sub = out_df[out_df["idiom"] == idiom].sort_values("lambda_years")
        scores_norm = _normalise_to_base(sub["score_paraphrase"].values)
        ax.plot(sub["lambda_years"], scores_norm,
                label=idiom, color=palette[i], linewidth=1.4, alpha=0.85)

    y_top = ax.get_ylim()[1] * 0.97
    _plot_robustness_markers(ax, checks, x_min, x_max, y_top)
    _add_legends(ax, checks, x_min, x_max)

    ax.set_xlabel("λ (projected year)", fontsize=12)
    ax.set_ylabel("Relative triviality score  (base = 1 at earliest λ)", fontsize=12)
    ax.set_title("Effect of temporal shift on triviality score — output space (indexed to 1)", fontsize=13)
    fig.tight_layout()
    fig.savefig(plots_dir / "output_intervention.png", dpi=150)
    plt.close(fig)
    logger.info("Saved → %s", plots_dir / "output_intervention.png")

    # --- Plot 2: Activation intervention curves (base-normalised, best layer) ---
    act_best = act_df[act_df["intervention_layer"] == best_layer].copy()
    x_min_act = float(act_best["lambda_years"].min()) if len(act_best) else x_min
    x_max_act = float(act_best["lambda_years"].max()) if len(act_best) else x_max

    fig, ax = plt.subplots(figsize=(14, 7))
    _draw_shared_annotations(ax, x_min_act, x_max_act)

    for i, idiom in enumerate(idioms):
        sub = act_best[act_best["idiom"] == idiom].sort_values("lambda_years")
        if len(sub) < 2:
            continue
        scores_norm = _normalise_to_base(sub["score_paraphrase"].values)
        ax.plot(sub["lambda_years"], scores_norm,
                label=idiom, color=palette[i], linewidth=1.4, alpha=0.85)

    y_top = ax.get_ylim()[1] * 0.97
    _plot_robustness_markers(ax, checks, x_min_act, x_max_act, y_top)
    _add_legends(ax, checks, x_min_act, x_max_act)

    ax.set_xlabel("λ (projected year)", fontsize=12)
    ax.set_ylabel("Relative triviality score  (base = 1 at earliest λ)", fontsize=12)
    ax.set_title(
        f"Effect of temporal shift on triviality score — activation space, Layer {best_layer} (indexed to 1)",
        fontsize=13,
    )
    fig.tight_layout()
    fig.savefig(plots_dir / "activation_intervention.png", dpi=150)
    plt.close(fig)
    logger.info("Saved → %s", plots_dir / "activation_intervention.png")

    # --- Plot 3: Layer sensitivity heatmap ---
    fig, ax = plt.subplots(figsize=(11, max(4, n_idioms * 0.45)))
    sns.heatmap(
        layer_mat,
        ax=ax,
        xticklabels=[f"L{i+1}" for i in range(n_layers)],
        yticklabels=idioms,
        center=0,
        cmap="RdBu_r",
        annot=True,
        fmt=".1e",
        linewidths=0.3,
        cbar_kws={"label": "Δscore_paraphrase / Δyear"},
    )
    ax.set_xlabel("BERT layer of injection", fontsize=12)
    ax.set_ylabel("")
    ax.set_title("Layer-wise sensitivity to temporal direction injection", fontsize=13)
    fig.tight_layout()
    fig.savefig(plots_dir / "layer_heatmap.png", dpi=150)
    plt.close(fig)
    logger.info("Saved → %s", plots_dir / "layer_heatmap.png")

    # --- Plot 4: Real value time series from BoE millennium data ---
    millennium_path = PROJECT_ROOT / paths["millennium_data"]
    _plot_real_value(millennium_path, cfg, plots_dir, x_min, x_max)

    print(f"\nPlots → {plots_dir}/")

    # --- Print robustness check reference table ---
    if checks:
        print(f"\n=== Robustness check events within λ range [{x_min:.0f}, {x_max:.0f}] ===\n")
        in_range = [c for c in checks if x_min <= c["year"] <= x_max]
        for c in sorted(in_range, key=lambda x: x["year"]):
            end = f"–{c['end_year']}" if "end_year" in c else ""
            print(f"  {c['year']}{end:6}  [{c['severity']:6}]  {c['event']}")


def main() -> None:
    with open(CONFIG_PATH) as fh:
        cfg = yaml.safe_load(fh)

    parser = argparse.ArgumentParser(
        description="Stage 05: Analyse intervention results."
    )
    parser.add_argument("--model", default="bert",
                        choices=list(cfg.get("models", {"bert": None}).keys()),
                        help="Which model's results to analyse (default: bert).")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--log-level", default="INFO",
                        choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    warnings.filterwarnings("ignore", category=FutureWarning)
    run(cfg, args.model, args.dry_run)


if __name__ == "__main__":
    main()
